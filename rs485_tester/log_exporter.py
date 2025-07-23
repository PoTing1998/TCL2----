import os
import re
import openpyxl
from openpyxl.styles import Font,PatternFill

class LogToExcelExporter:
    def __init__(self, log_file_path):
        self.log_file_path = log_file_path

    def export_to_excel(self, output_excel_path=None):
        if not os.path.exists(self.log_file_path):
            raise FileNotFoundError(f"找不到日誌檔案：{self.log_file_path}")

        with open(self.log_file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RS485 Log"
        ws.append(["時間", "方向", "內容"])
        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # 送出 = TX = 綠底；接收 = RX = 黃底 
        fill_tx = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 淺綠
        fill_rx = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 淺黃

        pattern = r'^\[(?P<timestamp>[\d\-\:\. ]+)\] \[(?P<direction>送出|接收)\] (?P<data>.+)$'

        for line in lines:
            line = line.strip()
            match = re.match(pattern, line)
            if match:
                ts = match.group("timestamp")
                direction = match.group("direction")
                data = match.group("data")
                ws.append([ts, direction, data])

                # 底色設定
                current_row = ws.max_row
                if direction == "送出":
                    for cell in ws[current_row]:
                        cell.fill = fill_tx
                elif direction == "接收":
                    for cell in ws[current_row]:
                        cell.fill = fill_rx
            
            else:
                alt_match = re.match(r'^\[(?P<timestamp>[\d\-\:\. ]+)\] (?P<message>.+)$', line)
                if alt_match:
                    ts = alt_match.group("timestamp")
                    message = alt_match.group("message")
                    ws.append([ts, "", message])
                    # 不設定顏色，表示未分類的行

        if output_excel_path is None:
            output_excel_path = os.path.splitext(self.log_file_path)[0] + ".xlsx"


                # 自動調整欄寬
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2  # 加 2 做緩衝
        wb.save(output_excel_path)
        print(f"✅ 匯出完成：{os.path.abspath(output_excel_path)}")
